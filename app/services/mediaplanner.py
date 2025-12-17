from __future__ import annotations

from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font
from datetime import date
from typing import List, Optional, Literal
import io
import logging
import os
import json
from pathlib import Path

import yaml
from pydantic import BaseModel, ValidationError
from typing import Dict


try:
    import httpx  # type: ignore
except ImportError:  # pragma: no cover
    httpx = None

try:
    from openpyxl import load_workbook  # type: ignore
    from openpyxl.cell.cell import MergedCell  # type: ignore
except ImportError:  # pragma: no cover
    load_workbook = None
    MergedCell = None  # type: ignore


log = logging.getLogger(__name__)

Mode = Literal["capacity", "budget"]
CAPACITY_CACHE: dict | None = None

class ParsedBrief(BaseModel):
    client: str
    brand: str
    period_start: date
    period_end: date

    formats: List[str]
    geo: List[str]
    ages: List[str]
    genders: List[str]
    interests: List[str]

    budget: Optional[float] = None
    frequency: Optional[float] = None
    goal: Optional[str] = None
    message: Optional[str] = None

    raw_text: Optional[str] = None


class PlanRow(BaseModel):
    format_name: str
    geo_name: str
    device: str
    period_start: date
    period_end: date
    mode: Mode

    capacity_imps: int
    plan_imps: int
    budget: float


class PlanMeta(BaseModel):
    agency: Optional[str] = None          # Агентство
    advertiser: Optional[str] = None      # Клиент/рекламодатель
    brand: Optional[str] = None           # Бренд
    campaign_name: Optional[str] = None   # Название кампании

    manager: Optional[str] = None         # Менеджер (сейлз/аккаунт)
    prepared_by: Optional[str] = None     # Кто подготовил медиаплан
    brief_date: Optional[date] = None     # Дата брифа

    # Техполя, если захотим тащить что-то из CampaignHub
    created_by: Optional[str] = None
    booking_id: Optional[int] = None

    # Параметры ЦА для ячейки "Позиция"
    ages: Optional[List[str]] = None
    genders: Optional[List[str]] = None
    interests: Optional[List[str]] = None



def _load_mp_cfg() -> dict:
    """
    Читает config.yaml и возвращает блок mediaplanner, если он есть.
    """
    cfg_path = Path("config.yaml")
    if not cfg_path.exists():
        return {}
    try:
        cfg = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
        mp_cfg = cfg.get("mediaplanner") or {}
        if not isinstance(mp_cfg, dict):
            return {}
        return mp_cfg
    except Exception as exc:  # pragma: no cover
        log.warning("Не удалось прочитать config.yaml: %s", exc)
        return {}


CAPACITY_CACHE: dict | None = None


def _parse_int(x: object) -> int:
    if x is None:
        return 0
    if isinstance(x, (int, float)):
        return int(x)
    if isinstance(x, str):
        s = x.replace(" ", "").replace("\u00a0", "")
        if not s or s.lower() == "none":
            return 0
        try:
            return int(float(s))
        except ValueError:
            return 0
    return 0


def _parse_float(x: object) -> float:
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    if isinstance(x, str):
        s = x.strip()
        if not s or s.lower() in ("none", "нет"):
            return 0.0
        s = s.replace(" ", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return 0.0
    return 0.0


def _parse_percent(x: object) -> float:
    v = _parse_float(x)
    # если нам дали "9.08" или "9.08%" -> трактуем как проценты
    if v > 1.0:
        return v / 100.0
    return v


def _get_capacity_data() -> dict:
    """
    Читает "Тестовая считалка емкости 2.xlsx" (лист "Считалка") и кэширует
    справочники:
      - formats: формат -> {capacity_imps, uniques, cpm, cpc}
      - geo: регион -> доля населения РФ
      - ages: возрастной сегмент -> доля населения
      - genders: "М"/"Ж" -> доля
      - interests: интерес -> доля населения
    """
    global CAPACITY_CACHE
    if CAPACITY_CACHE is not None:
        return CAPACITY_CACHE

    mp_cfg = _load_mp_cfg()
    capacity_path = (
        mp_cfg.get("capacity_path")
        or os.getenv("MEDIAPLANNER_CAPACITY_PATH")
        or "Тестовая считалка емкости 2.xlsx"
    )

    if load_workbook is None:
        raise RuntimeError(
            "Для работы с capacity нужен openpyxl: pip install openpyxl"
        )

    wb = load_workbook(capacity_path, data_only=True)
    ws = wb["Считалка"] if "Считалка" in wb.sheetnames else wb.active

    # -------- форматы --------
    formats: Dict[str, dict] = {}
    row = 2
    while row <= ws.max_row:
        name = ws.cell(row=row, column=1).value
        if not name:
            break
        name_str = str(name).strip()
        cap_imps = _parse_int(ws.cell(row=row, column=2).value)
        uniques = _parse_int(ws.cell(row=row, column=3).value)
        cpm = _parse_float(ws.cell(row=row, column=7).value)
        cpc = _parse_float(ws.cell(row=row, column=8).value)
        formats[name_str] = {
            "capacity_imps": cap_imps,
            "uniques": uniques,
            "cpm": cpm,
            "cpc": cpc,
        }
        row += 1

    # -------- гео (доля населения) --------
    geo: Dict[str, float] = {}
    row = 15
    while row <= ws.max_row:
        name = ws.cell(row=row, column=1).value
        if not name:
            row += 1
            continue
        name_str = str(name).strip()
        if name_str.startswith("Возраст"):
            break
        share = ws.cell(row=row, column=3).value
        geo[name_str] = _parse_percent(share)
        row += 1

    # -------- возраст --------
    ages: Dict[str, float] = {}
    for row in range(105, 111):
        name = ws.cell(row=row, column=1).value
        share = ws.cell(row=row, column=2).value
        if not name:
            continue
        ages[str(name).strip()] = _parse_percent(share)

    # -------- пол --------
    genders: Dict[str, float] = {}
    for row in range(113, 115):
        name = ws.cell(row=row, column=1).value
        share = ws.cell(row=row, column=2).value
        if not name:
            continue
        genders[str(name).strip()] = _parse_float(share)

    # -------- интересы --------
    interests: Dict[str, float] = {}
    for row in range(119, 180):
        name = ws.cell(row=row, column=1).value
        share = ws.cell(row=row, column=2).value
        if not name:
            continue
        interests[str(name).strip()] = _parse_float(share)

    CAPACITY_CACHE = {
        "formats": formats,
        "geo": geo,
        "ages": ages,
        "genders": genders,
        "interests": interests,
    }
    return CAPACITY_CACHE


def _compute_geo_factor(geo_specs: list[str], geo_table: Dict[str, float]) -> float:
    if not geo_specs:
        return 1.0
    norm = [str(g).strip().lower() for g in geo_specs if g]
    if any("рф" in g or "russia" in g for g in norm):
        return 1.0
    total = 0.0
    for spec in norm:
        for name, share in geo_table.items():
            n = name.lower()
            if spec in n or n in spec:
                total += share
    return total or 1.0


def _compute_age_factor(age_specs: list[str], ages_table: Dict[str, float]) -> float:
    if not age_specs:
        return 1.0
    norm_specs: list[str] = []
    for s in age_specs:
        if not s:
            continue
        s = str(s).strip()
        if s.lower() in ("any", "all", "все"):
            return 1.0
        norm_specs.append(s)
    if not norm_specs:
        return 1.0

    # сегменты из capacity
    segments: list[tuple[int, int, float]] = []
    for label, share in ages_table.items():
        text = str(label)
        if text.startswith("Младше"):
            hi = int("".join(ch for ch in text if ch.isdigit()) or "18")
            lo = 0
        elif "+" in text:
            lo = int("".join(ch for ch in text if ch.isdigit()) or "55")
            hi = 100
        else:
            parts = text.replace("–", "-").split("-")
            if len(parts) != 2:
                continue
            lo, hi = int(parts[0]), int(parts[1])
        segments.append((lo, hi, share))

    # объединяем интервалы из брифа (25-45 и т.п.)
    ranges: list[tuple[int, int]] = []
    for spec in norm_specs:
        s = spec.replace("–", "-").replace("+", "-100")
        parts = s.split("-")
        try:
            if len(parts) == 2 and parts[1]:
                lo, hi = int(parts[0]), int(parts[1])
            elif len(parts) == 1:
                lo, hi = int(parts[0]), 100
            else:
                continue
            ranges.append((lo, hi))
        except ValueError:
            continue
    if not ranges:
        return 1.0

    ranges.sort()
    merged: list[tuple[int, int]] = [ranges[0]]
    for lo, hi in ranges[1:]:
        last_lo, last_hi = merged[-1]
        if lo <= last_hi:
            merged[-1] = (last_lo, max(last_hi, hi))
        else:
            merged.append((lo, hi))

    def overlap(a: int, b: int, c: int, d: int) -> int:
        lo = max(a, c)
        hi = min(b, d)
        return max(0, hi - lo)

    total_share = 0.0
    for seg_lo, seg_hi, share in segments:
        seg_len = seg_hi - seg_lo
        if seg_len <= 0:
            continue
        covered = 0.0
        for lo, hi in merged:
            covered += overlap(seg_lo, seg_hi, lo, hi)
        covered = min(covered, seg_len)
        frac = covered / seg_len
        total_share += share * frac

    return total_share or 1.0


def _compute_gender_factor(gender_specs: list[str], genders_table: Dict[str, float]) -> float:
    if not gender_specs:
        return 1.0
    norm = [str(g).strip().lower() for g in gender_specs if g]
    if any(g in ("any", "all", "любой") for g in norm):
        return 1.0

    selected: set[str] = set()
    for g in norm:
        if g.startswith("m") or g.startswith("м"):
            selected.add("М")
        if g.startswith("f") or g.startswith("ж"):
            selected.add("Ж")

    if not selected:
        return 1.0

    return sum(genders_table.get(k, 0.0) for k in selected) or 1.0


def _compute_interest_factor(
    interest_specs: list[str], interests_table: Dict[str, float]
) -> float:
    if not interest_specs:
        return 1.0
    norm_specs = [str(i).strip().lower() for i in interest_specs if i]
    matches: list[float] = []
    for spec in norm_specs:
        for name, share in interests_table.items():
            n = name.lower()
            if spec in n or n in spec:
                matches.append(float(share))
    if matches:
        return max(matches)
    return 1.0


def _map_format_name(name: str, formats_table: Dict[str, dict]) -> Optional[str]:
    s = str(name).strip().lower()
    aliases = {
        "native stories": "Native Stories",
        "native story": "Native Stories",
        "stories": "Native Stories",
        "video": "Video",
        "skin": "Skin ",
        "skin ": "Skin ",
        "баннеры": "in-read",
        "banners": "in-read",
        "banner": "in-read",
    }
    if s in aliases:
        return aliases[s]
    for fmt in formats_table.keys():
        if s == fmt.lower():
            return fmt
    for fmt in formats_table.keys():
        if s in fmt.lower() or fmt.lower() in s:
            return fmt
    return None


async def parse_brief_with_llm(brief: str) -> ParsedBrief:
    """
    Отправляет бриф в Яндекс LLM и возвращает нормализованную структуру ParsedBrief.
    """
    mp_cfg = _load_mp_cfg()

    api_key = os.getenv("YANDEX_API_KEY") or mp_cfg.get("yandex_api_key")
    folder_id = os.getenv("YANDEX_FOLDER_ID") or mp_cfg.get("yandex_folder_id")
    model_name = os.getenv("YANDEX_MODEL") or mp_cfg.get("yandex_model") or "yandexgpt-lite"
    endpoint = (
        os.getenv("YANDEX_ENDPOINT")
        or mp_cfg.get("yandex_endpoint")
        or "https://llm.api.cloud.yandex.net/foundationModels/v1/completion"
    )

    if httpx is None:
        raise RuntimeError("Модуль httpx не установлен. Установи: pip install httpx")

    if not api_key or not folder_id:
        raise RuntimeError(
            "Не заданы ключи Яндекс LLM. "
            "Нужно прописать YANDEX_API_KEY и YANDEX_FOLDER_ID "
            "в переменных окружения или в config.yaml -> mediaplanner."
        )

    model_uri = f"gpt://{folder_id}/{model_name}"

    system_prompt = (
        "Ты медиапланер в digital-агентстве. "
        "На вход ты получаешь свободно написанный бриф на русском языке. "
        "Твоя задача — вернуть строго JSON без комментариев и текста вокруг, "
        "со следующими полями:\n"
        "{\n"
        '  \"client\": \"строка — название клиента/рекламодателя\",\n'
        '  \"brand\": \"строка — бренд\",\n'
        '  \"period_start\": \"YYYY-MM-DD\",\n'
        '  \"period_end\": \"YYYY-MM-DD\",\n'
        '  \"formats\": [\"banners\", \"video\", ...],\n'
        '  \"geo\": [\"РФ\", \"Москва\", ...],\n'
        '  \"ages\": [\"25-45\", \"25-55\", ...],\n'
        '  \"genders\": [\"M\",\"F\",\"any\"],\n'
        '  \"interests\": [\"семья\", \"дети\", ...],\n'
        '  \"budget\": число или null,\n'
        '  \"frequency\": число или null,\n'
        '  \"goal\": \"строка с целями и KPI\",\n'
        '  \"message\": \"основное рекламное сообщение\"\n'
        "}\n"
        "Если какое-то поле нельзя определить — ставь null или пустой список. "
        "НЕ пиши ничего кроме JSON."
    )

    payload = {
        "modelUri": model_uri,
        "completionOptions": {
            "stream": False,
            "temperature": 0.2,
            "maxTokens": "2000",
        },
        "messages": [
            {"role": "system", "text": system_prompt},
            {"role": "user", "text": brief},
        ],
    }

    headers = {
        "Authorization": f"Api-Key {api_key}",
        "Content-Type": "application/json",
    }

    log.info("Отправляем бриф в Яндекс LLM (modelUri=%s)", model_uri)

    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.post(endpoint, headers=headers, json=payload)

    if resp.status_code != 200:
        raise RuntimeError(f"Yandex LLM error {resp.status_code}: {resp.text}")

    data = resp.json()
    try:
        text = data["result"]["alternatives"][0]["message"]["text"]
    except Exception as exc:  # pragma: no cover
        raise RuntimeError(f"Неожиданный формат ответа Yandex LLM: {exc}; raw={data!r}") from exc

    def _extract_first_json(s: str) -> str:
        s = s.strip()

        # убираем код-блоки ``` и ```json
        if s.startswith("```"):
            s = s.replace("```json", "").replace("```", "").strip()

        # ищем первый объект { ... } с учётом вложенности и строк
        start = s.find("{")
        if start == -1:
            raise ValueError("В ответе LLM не найден символ '{'")

        depth = 0
        in_str = False
        prev = ""
        end = None

        for i, ch in enumerate(s[start:], start):
            if ch == '"' and prev != "\\":
                in_str = not in_str
            if not in_str:
                if ch == "{":
                    depth += 1
                elif ch == "}":
                    depth -= 1
                    if depth == 0:
                        end = i + 1
                        break
            prev = ch

        if end is None:
            raise ValueError("Не удалось найти конец JSON-объекта")

        return s[start:end]

    # --- единственный парсинг JSON ---
    try:
        json_str = _extract_first_json(text)
        raw_obj = json.loads(json_str)
    except Exception as exc:
        raise RuntimeError(f"LLM вернул невалидный JSON: {exc}; text={text!r}") from exc

    # дефолты / нормализация
    if not raw_obj.get("client"):
        raw_obj["client"] = "UNKNOWN"
    if not raw_obj.get("brand"):
        # если бренд не определён или null — подставляем клиента
        raw_obj["brand"] = raw_obj.get("client") or "UNKNOWN"

    raw_obj.setdefault("period_start", date.today().isoformat())
    raw_obj.setdefault("period_end", date.today().isoformat())
    raw_obj.setdefault("formats", ["banners"])
    raw_obj.setdefault("geo", ["РФ"])
    raw_obj.setdefault("ages", ["25-45"])
    raw_obj.setdefault("genders", ["any"])
    raw_obj.setdefault("interests", ["generic"])

    try:
        parsed = ParsedBrief(**raw_obj)
    except ValidationError as exc:
        raise RuntimeError(
            f"Не удалось привести JSON LLM к схеме ParsedBrief: {exc}"
        ) from exc

    parsed.raw_text = brief
    return parsed


def calculate_plan(
    parsed: ParsedBrief,
    mode: Mode,
    user_budget: Optional[float] = None,
) -> List[PlanRow]:
    """
    Первая реализация расчёта capacity / плана на базе
    "Тестовой считалки емкости":

    - capacity_imps = Емкость показов * GeoFactor * AgeFactor * GenderFactor * InterestFactor,
      дополнительно ограничиваем capacity уникальными * частота_РК.
    - В режиме "capacity" план = capacity, бюджет считается от CPM.
    - В режиме "budget" бюджет делится поровну между форматами,
      план считается от CPM, но не превышает capacity.
    """
    cap_data = _get_capacity_data()
    fmt_table = cap_data["formats"]
    geo_table = cap_data["geo"]
    ages_table = cap_data["ages"]
    genders_table = cap_data["genders"]
    interests_table = cap_data["interests"]

    rows: List[PlanRow] = []
    formats = parsed.formats or ["banners"]
    freq = parsed.frequency or 1.0

    total_budget = float(user_budget or parsed.budget or 0.0)
    per_format_budget = (
        total_budget / len(formats) if total_budget > 0 and formats else 0.0
    )

    for fmt in formats:
        fmt_key = _map_format_name(fmt, fmt_table)
        if not fmt_key:
            log.warning("calculate_plan: не нашли формат в capacity: %r", fmt)
            continue

        rec = fmt_table[fmt_key]

        geo_factor = _compute_geo_factor(parsed.geo or [], geo_table)
        age_factor = _compute_age_factor(parsed.ages or [], ages_table)
        gender_factor = _compute_gender_factor(parsed.genders or [], genders_table)
        interest_factor = _compute_interest_factor(parsed.interests or [], interests_table)

        base_cap = rec["capacity_imps"]
        if base_cap <= 0:
            log.warning("calculate_plan: у формата %s нет емкости показов", fmt_key)
            continue

        cap = int(
            base_cap
            * geo_factor
            * age_factor
            * gender_factor
            * interest_factor
        )

        uniq_cap = int(rec.get("uniques", 0) * freq) if rec.get("uniques") else 0
        if uniq_cap > 0:
            capacity_imps = min(cap, uniq_cap, base_cap)
        else:
            capacity_imps = min(cap, base_cap)

        cpm = float(rec.get("cpm") or 0.0)

        if mode == "capacity":
            plan_imps = capacity_imps
            budget = plan_imps / 1000.0 * cpm if cpm > 0 else 0.0
        else:  # mode == "budget"
            budget = per_format_budget
            if cpm > 0 and budget > 0:
                plan_imps = int(budget / cpm * 1000.0)
            else:
                plan_imps = 0
            plan_imps = min(plan_imps, capacity_imps)

        row = PlanRow(
            format_name=fmt_key,
            geo_name=(parsed.geo[0] if parsed.geo else "РФ"),
            device="all",
            period_start=parsed.period_start,
            period_end=parsed.period_end,
            mode=mode,
            capacity_imps=capacity_imps,
            plan_imps=plan_imps,
            budget=budget,
        )
        rows.append(row)

    if not rows:
        log.warning("calculate_plan: не удалось собрать ни одной строки плана")
    return rows


def build_campaign_name_from_parsed(parsed: ParsedBrief) -> str:
    month_label = parsed.period_start.strftime("%m.%Y")
    return f"{parsed.client}/{parsed.brand}/{month_label}/Innovation Lab"


def build_excel(
    meta: PlanMeta,
    rows,
    template_path: Optional[str] = None,
) -> io.BytesIO:
    """
    Заполняем шаблон Innovation Lab:
      - шапка (агентство, рекламодатель, бренд, название, период, менеджер, подготовил, дата)
      - строки медиаплана, начиная с 15-й строки:
        A: Площадка
        B: Позиция (Размещение + Соцдем + Интересы)
        C: Гео
        D: Место размещения
        E: Формат
        F: Девайс
        G: Тип ротации
        H: Объем размещения
        I: Ед. изм.
        J: Период
        K: CPM
        L: Стоимость
    """
    if load_workbook is None:
        raise RuntimeError(
            "Для генерации Excel нужен пакет openpyxl: pip install openpyxl"
        )

    if template_path is None:
        mp_cfg = _load_mp_cfg()
        template_path = (
            mp_cfg.get("template_path")
            or os.getenv("MEDIAPLANNER_TEMPLATE_PATH")
            or "media_plan_template.xlsx"
        )

    wb = load_workbook(template_path, data_only=False)

    try:
        ws = wb["МП"]
    except KeyError:
        ws = wb.active

    # --- Шапка (ячейки подогнаны под твой шаблон) ---

    # B2 — Агентство
    if meta.agency:
        ws["B2"] = meta.agency

    # B4 — Рекламодатель / Клиент
    if meta.advertiser:
        ws["B4"] = meta.advertiser

    # B5 — Бренд
    if meta.brand:
        ws["B5"] = meta.brand

    # B6 — Название кампании
    if meta.campaign_name:
        ws["B6"] = meta.campaign_name

    # B7 — Период размещения
    if rows:
        period_start = min(r.period_start for r in rows)
        period_end = max(r.period_end for r in rows)
        ws["B7"] = f"{period_start:%d.%m.%Y}-{period_end:%d.%m.%Y}"

    # B8 — Менеджер
    if meta.manager:
        ws["B8"] = meta.manager

    # B9 — Подготовил
    if meta.prepared_by:
        ws["B9"] = meta.prepared_by

    # B10 — Дата (брифа)
    if meta.brief_date:
        ws["B10"] = meta.brief_date

    # --- Строки медиаплана ---

    from openpyxl.styles import Alignment, Font  # type: ignore
    try:
        from openpyxl.cell.cell import MergedCell  # type: ignore
    except Exception:
        MergedCell = None  # type: ignore

    # В шаблоне заголовки на 13-й строке, 14-я часто мерджится — начинаем с 15-й
    row_index = 15
    if MergedCell is not None:
        # если вдруг 15-я тоже merged, пролистываем вниз
        while row_index <= ws.max_row and isinstance(
            ws.cell(row=row_index, column=1), MergedCell
        ):
            row_index += 1

    # текст для "Позиции"
    socdem_parts: list[str] = []
    if meta.genders:
        socdem_parts.append(", ".join(meta.genders))
    if meta.ages:
        socdem_parts.append(", ".join(meta.ages))
    socdem_text = ", ".join(socdem_parts) if socdem_parts else "—"

    interests_text = ", ".join(meta.interests or []) if meta.interests else "—"

    base_position_text = (
        "Размещение на страницах сайтов Innovation Lab\n"
        f"Соц дем: {socdem_text}\n"
        f"Интересы: {interests_text}"
    )

    for r in rows:
        # A (1) — Площадка
        ws.cell(row=row_index, column=1).value = "Innovation Lab"

        # B (2) — Позиция (многострочно, по вертикали по центру)
        pos_cell = ws.cell(row=row_index, column=2)
        pos_cell.value = base_position_text
        pos_cell.alignment = Alignment(wrap_text=True, vertical="center")
        pos_cell.font = Font(bold=True)

        # C (3) — Гео
        ws.cell(row=row_index, column=3).value = r.geo_name

        # D (4) — Место размещения
        ws.cell(row=row_index, column=4).value = (
            "Размещение на страницах сайтов, сетевое размещение"
        )

        # E (5) — Формат
        ws.cell(row=row_index, column=5).value = r.format_name

        # F (6) — Девайс
        ws.cell(row=row_index, column=6).value = "Кросс-девайс"  # или r.device

        # G (7) — Тип ротации
        ws.cell(row=row_index, column=7).value = "Динамика*"

        # H (8) — Объем размещения (показы)
        ws.cell(row=row_index, column=8).value = r.plan_imps

        # I (9) — Единица измерения
        ws.cell(row=row_index, column=9).value = "Показов"

        # J (10) — Период
        ws.cell(row=row_index, column=10).value = (
            f"{r.period_start:%d.%m.%Y}-{r.period_end:%d.%m.%Y}"
        )

        # K (11) — CPM (расчётно)
        if r.plan_imps and r.budget:
            cpm = r.budget * 1000.0 / float(r.plan_imps)
        else:
            cpm = 0.0
        ws.cell(row=row_index, column=11).value = cpm

        # L (12) — Стоимость до НДС
        ws.cell(row=row_index, column=12).value = r.budget

        # Остальные коэффициенты (M–S) пока заполняем нулями, если пустые
        for col in range(13, 20):
            cell = ws.cell(row=row_index, column=col)
            if cell.value is None:
                cell.value = 0

        row_index += 1

    # --- Возвращаем поток ---
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
