"""Parsers for XLSX/CSV attachments.

This module contains utility functions to turn downloaded files into plain
Python dictionaries suitable for insertion into the database.  Each parser
returns a list of daily records along with an optional period (date range)
inferred from the data itself.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple

from io import BytesIO
from openpyxl import load_workbook


DATE_PATTERNS = [
    # DD.MM.YYYY
    (re.compile(r"^(\d{2})\.(\d{2})\.(\d{4})$"), "%d.%m.%Y"),
    # YYYY-MM-DD
    (re.compile(r"^(\d{4})-(\d{2})-(\d{2})$"), "%Y-%m-%d"),
]


def parse_date(value: str) -> Optional[date]:
    """Parse a date from common formats into a date object.

    Returns None if parsing fails.
    """
    if not value:
        return None
    s = str(value).strip()
    for regex, fmt in DATE_PATTERNS:
        if regex.match(s):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
    return None


def normalize_column_name(name: str) -> str:
    """Return a lowercased key suitable for mapping to canonical fields."""
    return re.sub(r"\s+", " ", name.strip().lower())


def parse_system_xlsx(data: bytes) -> Tuple[List[Dict], Optional[date], Optional[date]]:
    """Parse an XLSX file exported from your advertising system.

    The parser looks for common Russian/English column names and maps them to
    canonical fields: date, impressions, clicks, spend, reach, frequency,
    CTR and various viewability metrics.  It returns a list of dicts and
    the inferred (min_date, max_date) range.  Unknown columns are ignored.
    The implementation uses openpyxl directly to avoid a dependency on pandas.
    """
    wb = load_workbook(filename=BytesIO(data), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    # Extract header row
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], None, None
    # Map columns to canonical keys
    mapping = {
        "день": "date",
        "day": "date",
        "date": "date",
        "показы": "impressions",
        "impressions": "impressions",
        "переходы": "clicks",
        "clicks": "clicks",
        "бюджет": "spend",
        "расход": "spend",
        "spend": "spend",
        "cost": "spend",
        "охват": "reach",
        "reach": "reach",
        "частота": "frequency",
        "frequency": "frequency",
        "ctr": "ctr",
        "просмотр 1/4": "view_25",
        "просмотр 1/2": "view_50",
        "просмотр 3/4": "view_75",
        "досмотр": "view_100",
        "vtr": "vtr",
    }
    # Build index mapping from column index to canonical key
    idx_to_key: Dict[int, str] = {}
    for idx, col_name in enumerate(header):
        if col_name is None:
            continue
        key = mapping.get(normalize_column_name(str(col_name)), None)
        if key:
            idx_to_key[idx] = key
    canonical_records: List[Dict] = []
    min_date: Optional[date] = None
    max_date: Optional[date] = None
    for row in rows_iter:
        record: Dict = {}
        for idx, value in enumerate(row):
            key = idx_to_key.get(idx)
            if not key:
                continue
            if key == "date":
                dt = None
                if value is not None:
                    if isinstance(value, datetime):
                        dt = value.date()
                    else:
                        dt = parse_date(str(value))
                record["date"] = dt
                if dt:
                    if min_date is None or dt < min_date:
                        min_date = dt
                    if max_date is None or dt > max_date:
                        max_date = dt
            else:
                if value is None:
                    record[key] = None
                else:
                    s = str(value).replace(" ", "").replace(",", ".")
                    if key in ("impressions", "clicks", "view_25", "view_50", "view_75", "view_100", "reach"):
                        try:
                            record[key] = int(float(s))
                        except Exception:
                            record[key] = None
                    elif key in ("spend", "frequency", "ctr", "vtr"):
                        try:
                            record[key] = float(s)
                        except Exception:
                            record[key] = None
        if record.get("date"):
            canonical_records.append(record)
    return canonical_records, min_date, max_date


def parse_metrica_xlsx(data: bytes) -> Tuple[List[Dict], Optional[date], Optional[date]]:
    """Parse an XLSX file exported from Yandex Metrica via mail.

    Metrica reports often contain several header lines describing filters
    and attribution models.  The parser skips non‑tabular rows and then
    looks for a `Date` or `Дата` column and a handful of metric columns.
    Returns a list of records and the date range.
    The implementation uses openpyxl to avoid pandas and gracefully
    handles varying header positions.
    """
    wb = load_workbook(filename=BytesIO(data), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    # Find the first row that contains 'Дата', 'Date' or 'UTM'
    header_row_idx = None
    rows = list(ws.iter_rows(values_only=True))
    for idx, row in enumerate(rows):
        combined = " ".join(str(x) for x in row if x is not None)
        lower = combined.lower()
        if "дата" in lower or "date" in lower or "utm" in lower:
            header_row_idx = idx
            break
    if header_row_idx is None:
        return [], None, None
    header = rows[header_row_idx]
    # Build mapping from column index to canonical key
    mapping = {
        "дата": "date",
        "date": "date",
        "day": "date",
        "utm campaign": "utm_campaign",
        "utm source": "utm_source",
        "utm content": "utm_content",
        "визиты": "visits",
        "visits": "visits",
        "посетители": "visitors",
        "users": "visitors",
        "отказы": "bounces",
        "bounce rate": "bounce_rate",
        "показатель отказов": "bounce_rate",
        "глубина просмотра": "depth",
        "average page depth": "depth",
        "время на сайте": "time_on_site",
        "average visit duration": "time_on_site",
        "конверсии": "conversions",
        "conversions": "conversions",
    }
    idx_to_key: Dict[int, str] = {}
    for i, col in enumerate(header):
        if col is None:
            continue
        key = mapping.get(normalize_column_name(str(col)), None)
        if key:
            idx_to_key[i] = key
    records: List[Dict] = []
    min_date: Optional[date] = None
    max_date: Optional[date] = None
    # Iterate over rows after the header
    for row in rows[header_row_idx + 1:]:
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        # Skip summary rows (Итого ...)
        if any(isinstance(val, str) and val.strip().lower().startswith("итого") for val in row):
            continue
        record: Dict = {}
        for idx, val in enumerate(row):
            key = idx_to_key.get(idx)
            if not key:
                continue
            if key == "date":
                dt = None
                if val is not None:
                    if isinstance(val, datetime):
                        dt = val.date()
                    else:
                        dt = parse_date(str(val))
                record["date"] = dt
                if dt:
                    if min_date is None or dt < min_date:
                        min_date = dt
                    if max_date is None or dt > max_date:
                        max_date = dt
            else:
                if val is None or (isinstance(val, float) and str(val) == 'nan'):
                    record[key] = None
                else:
                    s = str(val).replace(" ", "").replace(",", ".")
                    if key in ("visits", "visitors", "bounces", "conversions"):
                        try:
                            record[key] = int(float(s))
                        except Exception:
                            record[key] = None
                    elif key in ("depth", "bounce_rate"):
                        try:
                            record[key] = float(s)
                        except Exception:
                            record[key] = None
                    elif key == "time_on_site":
                        # convert HH:MM:SS or mm:ss to seconds
                        try:
                            parts = [float(p) for p in s.split(":")]
                            seconds = 0
                            if len(parts) == 3:
                                seconds = parts[0] * 3600 + parts[1] * 60 + parts[2]
                            elif len(parts) == 2:
                                seconds = parts[0] * 60 + parts[1]
                            record[key] = seconds
                        except Exception:
                            record[key] = None
        if record.get("date"):
            records.append(record)
    return records, min_date, max_date
